import os
import sys

sys.path.append('/home/dylan/catkin_ws/src/mask_rcnn_ros/src/')

import json
import numpy as np
import json
import skimage
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from tensorflow.python.client import device_lib

from Mask_RCNN.mrcnn import visualize as viz
from Mask_RCNN.scripts.visualize_cv2 import model, class_names

# These are the imports from other files
from evaluate import *

print(device_lib.list_local_devices())

# You need to have bboxResults initialised as a dictionary of numpy arrays.
# bboxCFList is a list of numpy arrays in the same order as bboxResults

# Format your bbox and confidence level results (if needed) and append them in these variables
real_test_dir = '/home/dylan/catkin_ws/src/mask_rcnn_ros/src/Mask_RCNN/datasets/droneDetection/real_test/image_10m/'
export_path = "/home/dylan/catkin_ws/src/mask_rcnn_ros/src/Mask_RCNN/exports/results.xlsx"
labelFilePath = "/home/dylan/catkin_ws/src/mask_rcnn_ros/src/Mask_RCNN/datasets/droneDetection/real_test/label.json"

image_paths = []
for filename in sorted(os.listdir(real_test_dir)):
    print(filename)
    if os.path.splitext(filename)[1].lower() in ['.png', '.jpg', '.jpeg']:
        image_paths.append(os.path.join(real_test_dir, filename))
        
bboxResults = {}
bboxCFList = []
        
for image_path in image_paths:
    img = skimage.io.imread(image_path)
    img_arr = np.array(img)
    results = model.detect([img_arr], verbose=1)
    r = results[0]
    bboxCFList.append(r['scores'])
    bboxResults[image_path.replace(real_test_dir, "")] = r['rois'].tolist()
    # viz.display_instances(img, r['rois'], r['masks'], r['class_ids'], 
    #                         class_names, r['scores'], figsize=(15,15))
    
# print(f"{image_paths}\n")
print(f"{bboxResults}\n")
# print(f"{bboxCFList}\n")

# First convert dictionary of lists into a dataframe, with the image name as index
df = pd.DataFrame.from_dict(bboxResults,orient='index', columns=['BBox Array'])

# Convert list of Confidence Level numpy array into pandas Series and append into the dataframe
cfdf = pd.Series(bboxCFList)
df = df.assign(CF=cfdf.values)
print(df)


try: 
    df.to_excel(export_path, header=True)
except PermissionError:
    print("There is already an existing file. You should remove it if you want to overwrite the file.")


# First transfer raw bounding box values and confidence levels of each image into an excel sheet
transfer_bbox(export_path,labelFilePath)

# Set the column names of each column used
set_header(export_path)

# Calculate and store each evaluation matrix
totalImg = calculate_total_images(export_path, bboxResults)
fNegCount = calculate_false_negatives(export_path, bboxResults)
fPosCount = calculate_false_positives(export_path, bboxResults)
passingRate = calculate_passing_rate(export_path, bboxResults)
average_cf = calculate_avg_cf(export_path, bboxResults)

# Finally, clean up the excel sheet for easy reading
clean_excel(export_path)